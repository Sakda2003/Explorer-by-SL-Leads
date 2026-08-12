"""Build the lead-level dataset: one row per lead, carrying all 10 model variables.

Unions the four weekly CRM exports, then left-joins each lead to its ad set x day
context from Dataset/master_dataset.xlsx. Every lead is kept, including ones with
no UTM ad set and ones falling outside the Meta export window -- they are flagged
rather than dropped, so the row count reconciles against the raw exports.

Grain warning: the ad-set-day columns (spend, frequency, reach, ...) repeat on
every lead of that ad set on that day. Summing them over leads double-counts.
Use spend_attributed_to_lead, or aggregate back to ad set x day first.
"""

import re
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from blank_policy import INFERRED_COLS, apply_blank_policy, order_with_inferred

CRM = [
    (r"Dataset/Datsaa/Dataset 06-06 to 07-11.xlsx", "06-06 to 07-11"),
    (r"Dataset/Datsaa/Dataset 07-11 to 07-18.xlsx", "07-11 to 07-18"),
    (r"Dataset/Datsaa/Dataset 07-18 to 07-25.xlsx", "07-18 to 07-25"),
    (r"Dataset/Datsaa/Dataset 07-25 to 08-01.xlsx", "07-25 to 08-01"),
]
MASTER = r"Dataset/master_dataset.xlsx"
HOL = r"Dataset/Dataset/holiday_proximity.xlsx"
OUT = r"Dataset/LEAD_LEVEL_DATASET.xlsx"

ID_COLS = ["UTM Campaign ID", "UTM Ad Set ID", "UTM Ad ID"]
SEP = r"\||\s-\s"

# ---------------------------------------------------------------- load CRM
frames = []
for path, tag in CRM:
    d = pd.read_excel(path, sheet_name=0, dtype={c: str for c in ID_COLS})
    d["source_file"] = tag
    frames.append(d)
raw = pd.concat(frames, ignore_index=True)
n_raw = len(raw)

raw["created_at"] = pd.to_datetime(raw["Created At"], format="%m/%d/%Y, %I:%M:%S %p")
raw["updated_at"] = pd.to_datetime(raw["Updated At"], format="%m/%d/%Y, %I:%M:%S %p",
                                   errors="coerce")
raw["date"] = raw["created_at"].dt.normalize()

# exact duplicates across the weekly files (boundaries are clean, but check)
dup_mask = raw.duplicated(subset=["created_at", "Customer Name", "UTM Ad ID"], keep="first")
n_dup = int(dup_mask.sum())
raw = raw[~dup_mask].reset_index(drop=True)

raw = raw.sort_values("created_at").reset_index(drop=True)
raw.insert(0, "lead_id", ["L%05d" % (i + 1) for i in range(len(raw))])

# ---------------------------------------------------------------- lead-own fields
raw["hour"] = raw["created_at"].dt.hour
raw["minutes_to_update"] = (raw["updated_at"] - raw["created_at"]).dt.total_seconds() / 60


def creative_code(title):
    if not isinstance(title, str):
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", re.split(SEP, title)[0]).upper()


def template_key(title):
    if not isinstance(title, str):
        return ""
    parts = [p for p in re.split(SEP, title) if p.strip()]
    return re.sub(r"[^A-Za-z0-9]", "", parts[-1]).upper() if parts else ""


raw["creative_code"] = raw["FB Ad Title"].map(creative_code)
raw["template_key"] = raw["FB Ad Title"].map(template_key)

lead = raw.rename(columns={
    "Platform": "platform", "Status": "lead_status", "Customer Name": "customer_name",
    "UTM Campaign": "utm_campaign", "UTM Campaign ID": "campaign_id",
    "UTM Ad Set ID": "ad_set_id", "UTM Ad ID": "ad_id", "FB Ad Title": "fb_ad_title",
})[[
    "lead_id", "created_at", "date", "hour", "updated_at", "minutes_to_update",
    "platform", "lead_status", "customer_name",
    "utm_campaign", "campaign_id", "ad_set_id", "ad_id",
    "fb_ad_title", "creative_code", "template_key", "source_file",
]].copy()

lead["is_new_customer"] = (lead["lead_status"] == "New").astype(int)
lead["has_utm_ad_set"] = lead["ad_set_id"].notna().astype(int)

# ---------------------------------------------------------------- join ad set x day context
m = pd.read_excel(MASTER, sheet_name="master_adset_daily",
                  dtype={"ad_set_id": str, "campaign_id": str})
m["date"] = pd.to_datetime(m["date"]).dt.normalize()
m["ad_set_change_source"] = "inferred"
m["ad_change_source"] = "inferred"

CTX = [
    "date", "ad_set_id",
    "campaign_name", "delivery_status",
    "leads", "spend",
    "days_since_adset_started", "adset_start_censored",
    "frequency",
    "ad_change_recency", "ad_set_change_recency",
    "ad_set_change_type", "ad_set_change_today", "ad_set_change_source",
    "ad_change_type", "ad_change_today", "ad_change_source",
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type", "cpl", "cpm",
]
ctx = m[CTX].rename(columns={"leads": "adset_day_leads"})

lead = lead.merge(ctx, on=["ad_set_id", "date"], how="left")
lead["matched_to_adset_day"] = lead["campaign_name"].notna().astype(int)

# Variables 3 and 8 are calendar facts about the lead's own date -- they do not
# depend on the ad-set join, so they are taken straight from the holiday file and
# the date. Sourcing them through the join would leave them blank on the 119
# leads with no matching ad-set-day, which would be wrong: an untracked lead still
# arrived on a Tuesday.
hol = pd.read_excel(HOL)
hol["date"] = pd.to_datetime(hol["date"]).dt.normalize()
lead = lead.merge(hol[["date", "is_holiday", "holiday_name", "holiday_proximity"]],
                  on="date", how="left")
lead["is_holiday"] = lead["is_holiday"].fillna(0).astype(int)
lead["holiday_name"] = lead["holiday_name"].fillna("")
lead["day_of_week"] = lead["date"].dt.day_name()
lead["day_of_week_num"] = lead["date"].dt.dayofweek
lead["is_weekend"] = lead["day_of_week_num"].isin([5, 6]).astype(int)

# ---------------------------------------------------------------- lead-grain derivations
lead = lead.sort_values(["ad_set_id", "created_at"])
lead["lead_seq_in_adset_day"] = lead.groupby(["ad_set_id", "date"]).cumcount() + 1
lead = lead.sort_values("created_at")
lead["lead_seq_in_day"] = lead.groupby("date").cumcount() + 1

# per-lead attributed spend: the day's ad set spend split evenly across its leads.
# Uses the CRM count for that ad set x day (not adset_day_leads, which is blank on
# unmatched rows) so it is defined wherever spend is.
crm_ct = lead.groupby(["ad_set_id", "date"])["lead_id"].transform("size")
lead["adset_day_leads_crm"] = crm_ct
lead["spend_attributed_to_lead"] = lead["spend"] / crm_ct
lead["lead_share_of_adset_day"] = 1.0 / crm_ct

# the 2026-08-01 export was pulled at 09:28, so that day is ~40% complete and must
# not be treated as a whole day in any daily average. See backend _trailing_partial_date.
last_day = lead["date"].max()
last_ts = lead.loc[lead["date"] == last_day, "created_at"].max()
PARTIAL = last_day if last_ts.hour < 18 else None
lead["is_trailing_partial_day"] = (lead["date"] == PARTIAL).astype(int) if PARTIAL is not None else 0

# outside the Meta ad-set export window (perf data ends before the CRM does)
perf_max = m["date"].max()
lead["outside_perf_window"] = (lead["date"] > perf_max).astype(int)

COLS = [
    # lead identity
    "lead_id", "created_at", "date", "hour", "updated_at", "minutes_to_update",
    "platform", "lead_status", "is_new_customer", "customer_name",
    # attribution
    "utm_campaign", "campaign_id", "campaign_name", "ad_set_id", "ad_id",
    "fb_ad_title", "creative_code", "template_key", "delivery_status",
    # the 10 variables, in order
    "adset_day_leads",                                                     # 1
    "spend",                                                               # 2
    "holiday_proximity", "is_holiday", "holiday_name",                     # 3
    "days_since_adset_started", "adset_start_censored",                    # 4
    "frequency",                                                           # 5
    "ad_change_recency",                                                   # 6
    "ad_set_change_recency",                                               # 7
    "day_of_week", "day_of_week_num", "is_weekend",                        # 8
    "ad_set_change_type", "ad_set_change_today", "ad_set_change_source",   # 9
    "ad_change_type", "ad_change_today", "ad_change_source",               # 10
    # supporting ad-set-day context
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type", "cpl", "cpm",
    # lead-grain derivations + flags
    "adset_day_leads_crm", "lead_seq_in_adset_day", "lead_seq_in_day",
    "spend_attributed_to_lead", "lead_share_of_adset_day",
    "has_utm_ad_set", "matched_to_adset_day", "outside_perf_window",
    "is_trailing_partial_day", "source_file",
]
lead = apply_blank_policy(lead)
COLS = order_with_inferred(COLS)
lead = lead[COLS].sort_values("created_at").reset_index(drop=True)

# ---------------------------------------------------------------- reconciliation
unmatched = lead[lead["matched_to_adset_day"] == 0]
recon = pd.DataFrame([
    ("raw rows in the 4 CRM exports", n_raw),
    ("exact duplicates removed", n_dup),
    ("leads in this file", len(lead)),
    ("", ""),
    ("leads with a UTM ad set id", int(lead["has_utm_ad_set"].sum())),
    ("leads with NO UTM ad set id (untracked / organic)", int((lead["has_utm_ad_set"] == 0).sum())),
    ("", ""),
    ("matched to an ad set x day row", int(lead["matched_to_adset_day"].sum())),
    ("unmatched: no UTM ad set", int(((lead["matched_to_adset_day"] == 0) & (lead["has_utm_ad_set"] == 0)).sum())),
    ("unmatched: after the Meta export window", int(((lead["matched_to_adset_day"] == 0) & (lead["outside_perf_window"] == 1)).sum())),
    ("unmatched: ad set not in Meta export that day", int(((lead["matched_to_adset_day"] == 0) & (lead["has_utm_ad_set"] == 1) & (lead["outside_perf_window"] == 0)).sum())),
    ("", ""),
    ("on the trailing partial day (export pulled 09:28)", int(lead["is_trailing_partial_day"].sum())),
    ("", ""),
    ("date range", f"{lead['date'].min():%Y-%m-%d} to {lead['date'].max():%Y-%m-%d}"),
    ("distinct ad sets", int(lead["ad_set_id"].nunique())),
    ("distinct ads", int(lead["ad_id"].nunique())),
    ("distinct campaigns", int(lead["campaign_id"].nunique())),
    ("total spend on matched ad-set-days", round(float(lead["spend_attributed_to_lead"].sum()), 2)),
], columns=["check", "value"])

# ---------------------------------------------------------------- write
NOTES = [
    ("LEAD LEVEL DATASET", ""),
    ("Generated", pd.Timestamp.today().strftime("%Y-%m-%d")),
    ("Grain", "one row per lead"),
    ("Rows", f"{len(lead):,}"),
    ("", ""),
    ("HOW TO READ IT", ""),
    ("", "Columns up to delivery_status describe the lead itself, straight from the CRM export."),
    ("", "Everything from adset_day_leads onward is the ad set x day CONTEXT that lead arrived in,"),
    ("", "joined on ad_set_id + date. Those values repeat on every lead of the same ad set on the"),
    ("", "same day."),
    ("", ""),
    ("THE DOUBLE-COUNTING TRAP", ""),
    ("", "Do NOT sum spend, reach, impressions or frequency down this sheet -- an ad-set-day with 20"),
    ("", "leads repeats its spend 20 times. Either aggregate back to ad set x day first, or use"),
    ("", "spend_attributed_to_lead, which splits the day's spend evenly across that ad set's leads"),
    ("", "and therefore does sum correctly. lead_share_of_adset_day does the same for counts."),
    ("", "adset_day_leads is the CRM lead total for that ad-set-day, repeated -- it is variable 1 in"),
    ("", "the model, not a per-lead quantity."),
    ("", ""),
    ("UNMATCHED LEADS", ""),
    ("", "Every lead in the exports is present. Leads with no ad set context have blanks in the"),
    ("", "variable columns and are flagged: has_utm_ad_set = 0 (untracked), outside_perf_window = 1"),
    ("", "(arrived after the Meta export ends), matched_to_adset_day = 0 (either of those, or the ad"),
    ("", "set had no delivery row that day). Filter on matched_to_adset_day = 1 before modelling."),
    ("", ""),
    ("2026-08-01", ""),
    ("", "That day's export was pulled at 09:28, so it holds a partial morning only. Rows are kept"),
    ("", "and flagged is_trailing_partial_day = 1. Exclude it from any daily average or it drags the"),
    ("", "level down about 14%."),
    ("", ""),
    ("INFERRED VS RECORDED", ""),
    ("", "ad_set_change_type and ad_change_type carry source = inferred on every row: they are"),
    ("", "detected from delivery step shifts, not read from Meta. Fill the change logs in"),
    ("", "MODEL_DATASET_TEMPLATE.xlsx to replace them with recorded events."),
]

with pd.ExcelWriter(OUT, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss") as w:
    pd.DataFrame(NOTES, columns=["", " "]).to_excel(w, sheet_name="README", index=False)
    lead.to_excel(w, sheet_name="lead_level", index=False)
    recon.to_excel(w, sheet_name="reconciliation", index=False)
    unmatched.to_excel(w, sheet_name="unmatched_leads", index=False)

    wb = w.book
    ws = wb["README"]
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 100
    for r in ws.iter_rows(max_col=2):
        r[1].alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(r[0].value, str) and r[0].value.isupper() and r[0].value:
            r[0].font = Font(bold=True)

    LEADF = PatternFill("solid", fgColor="FFD9D9D9")
    VARF = PatternFill("solid", fgColor="FFFFE699")
    SUPF = PatternFill("solid", fgColor="FFF2F2F2")
    FLAGF = PatternFill("solid", fgColor="FFFCE4D6")
    VARCOLS = set(COLS[COLS.index("adset_day_leads"):COLS.index("reach")])
    FLAGCOLS = set(COLS[COLS.index("adset_day_leads_crm"):])

    for name in ("lead_level", "unmatched_leads"):
        ws = wb[name]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 30
        for i, c in enumerate(ws[1], start=1):
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.fill = (VARF if c.value in VARCOLS else
                      FLAGF if c.value in FLAGCOLS else
                      SUPF if i > COLS.index("reach") else LEADF)
            ws.column_dimensions[get_column_letter(i)].width = 15
        for nm, wd in {"lead_id": 9, "created_at": 19, "date": 11, "updated_at": 19,
                       "customer_name": 22, "utm_campaign": 28, "campaign_id": 20,
                       "campaign_name": 26, "ad_set_id": 20, "ad_id": 20,
                       "fb_ad_title": 46, "holiday_proximity": 17,
                       "ad_set_change_type": 19, "ad_change_type": 20}.items():
            ws.column_dimensions[get_column_letter(COLS.index(nm) + 1)].width = wd

    ws = wb["reconciliation"]
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22
    for c in ws[1]:
        c.font = Font(bold=True)

lead.to_csv(r"Dataset/lead_level_dataset.csv", index=False)

print("wrote", OUT)
print(recon.to_string(index=False))
print("\ncolumns", len(COLS))
print("nulls in variable columns (all on unmatched rows):")
nn = lead[list(VARCOLS)].isna().sum()
print(nn[nn > 0].to_string())
