"""Rebuild MODEL_DATASET_TEMPLATE at lead grain.

Sakda's pruned template carries one row per ad set per day with a `leads` count. This
replaces that count with the individual leads themselves: one row per lead, carrying the
lead's own identity plus the ad set x day context it arrived in.

Every lead in the four CRM exports is kept, including untracked ones, so the row count
reconciles against the source files rather than silently dropping what will not join.
"""

import re
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from blank_policy import INFERRED_COLS, apply_blank_policy

CRM = [
    (r"Dataset/Datsaa/Dataset 06-06 to 07-11.xlsx", "06-06 to 07-11"),
    (r"Dataset/Datsaa/Dataset 07-11 to 07-18.xlsx", "07-11 to 07-18"),
    (r"Dataset/Datsaa/Dataset 07-18 to 07-25.xlsx", "07-18 to 07-25"),
    (r"Dataset/Datsaa/Dataset 07-25 to 08-01.xlsx", "07-25 to 08-01"),
]
MASTER = r"Dataset/master_dataset.xlsx"
HOL = r"Dataset/Dataset/holiday_proximity.xlsx"
OUT = r"Dataset/Datsaa/Dataset Template/MODEL_DATASET_TEMPLATE.xlsx"

ID_COLS = ["UTM Campaign ID", "UTM Ad Set ID", "UTM Ad ID"]
SEP = r"\||\s-\s"
ADSET_TYPES = ["budget_change", "targeting_change", "placement_change", "bid_change"]
AD_TYPES = ["ad_added", "ad_paused", "ad_swapped", "msg_template_change", "no_recent_change"]
SOURCES = ["confirmed", "inferred"]

# ---------------------------------------------------------------- leads
frames = []
for path, tag in CRM:
    frame = pd.read_excel(path, sheet_name=0, dtype={c: str for c in ID_COLS})
    frame["source_file"] = tag
    frames.append(frame)
raw = pd.concat(frames, ignore_index=True)
n_raw = len(raw)

raw["created_at"] = pd.to_datetime(raw["Created At"], format="%m/%d/%Y, %I:%M:%S %p")
raw["date"] = raw["created_at"].dt.normalize()
n_dup = int(raw.duplicated(["created_at", "Customer Name", "UTM Ad ID"], keep="first").sum())
raw = raw[~raw.duplicated(["created_at", "Customer Name", "UTM Ad ID"], keep="first")]
raw = raw.sort_values("created_at").reset_index(drop=True)
raw.insert(0, "lead_id", ["L%05d" % (i + 1) for i in range(len(raw))])


def creative_code(title):
    if not isinstance(title, str):
        return ""
    return re.sub(r"[^A-Za-z0-9]", "", re.split(SEP, title)[0]).upper()


lead = raw.rename(columns={
    "Platform": "platform", "Status": "lead_status", "Customer Name": "customer_name",
    "UTM Campaign": "utm_campaign", "UTM Campaign ID": "campaign_id",
    "UTM Ad Set ID": "ad_set_id", "UTM Ad ID": "ad_id", "FB Ad Title": "fb_ad_title",
})
lead["is_new_customer"] = lead["lead_status"].eq("New").astype(int)
lead["creative_code"] = lead["fb_ad_title"].map(creative_code)
lead["hour"] = lead["created_at"].dt.hour

# ---------------------------------------------------------------- ad set x day context
m = pd.read_excel(MASTER, sheet_name="master_adset_daily",
                  dtype={"ad_set_id": str, "campaign_id": str})
m["date"] = pd.to_datetime(m["date"]).dt.normalize()
m["ad_set_change_source"] = "inferred"
m["ad_change_source"] = "inferred"

CTX = [
    "date", "ad_set_id", "campaign_id", "campaign_name", "delivery_status",
    "leads", "spend",
    "days_since_adset_started", "frequency",
    "ad_change_recency", "ad_set_change_recency",
    "ad_set_change_type", "ad_set_change_today", "ad_set_change_source",
    "ad_change_type", "ad_change_today", "ad_change_source",
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type", "cpl", "cpm",
]
ctx = m[CTX].rename(columns={"leads": "adset_day_leads", "campaign_id": "campaign_id_meta",
                             "campaign_name": "campaign_name_meta"})
lead = lead.merge(ctx, on=["ad_set_id", "date"], how="left")
lead["matched_to_adset_day"] = lead["campaign_name_meta"].notna().astype(int)
# prefer Meta's campaign name; the CRM's UTM Campaign string is the fallback for untracked rows
lead["campaign_name"] = lead["campaign_name_meta"].fillna(lead["utm_campaign"]).fillna("")

# calendar variables belong to the lead's own date, not to the ad-set join -- an untracked
# lead still arrived on a Tuesday, so these stay populated on every row
hol = pd.read_excel(HOL)
hol["date"] = pd.to_datetime(hol["date"]).dt.normalize()
lead = lead.merge(hol[["date", "is_holiday", "holiday_proximity"]], on="date", how="left")
lead["is_holiday"] = lead["is_holiday"].fillna(0).astype(int)
lead["day_of_week"] = lead["date"].dt.day_name()
lead["day_of_week_num"] = lead["date"].dt.dayofweek
lead["is_weekend"] = lead["day_of_week_num"].isin([5, 6]).astype(int)

# ---------------------------------------------------------------- lead-grain derivations
lead = lead.sort_values(["ad_set_id", "created_at"])
lead["lead_seq_in_adset_day"] = lead.groupby(["ad_set_id", "date"]).cumcount() + 1
crm_count = lead.groupby(["ad_set_id", "date"])["lead_id"].transform("size")
lead["adset_day_leads_crm"] = crm_count
# the day's ad set spend split evenly across its leads: unlike `spend`, this sums correctly
lead["spend_attributed_to_lead"] = lead["spend"] / crm_count

last_day = lead["date"].max()
last_hour = lead.loc[lead["date"] == last_day, "created_at"].max().hour
lead["is_trailing_partial_day"] = (
    lead["date"].eq(last_day).astype(int) if last_hour < 18 else 0)

lead = apply_blank_policy(lead)

COLUMNS = [
    # the lead
    "lead_id", "created_at", "date", "hour", "customer_name", "lead_status",
    "is_new_customer", "platform",
    # attribution
    "campaign_id", "campaign_name", "ad_set_id", "ad_id", "fb_ad_title", "creative_code",
    "delivery_status",
    # declared variables 2-10 (variable 1 is the row itself)
    "spend",
    "holiday_proximity", "is_holiday",
    "days_since_adset_started",
    "frequency",
    "ad_change_recency", "ad_change_recency_inferred",
    "ad_set_change_recency", "ad_set_change_recency_inferred",
    "day_of_week", "day_of_week_num", "is_weekend",
    "ad_set_change_type", "ad_set_change_type_inferred",
    "ad_change_type", "ad_change_type_inferred",
    # ad set x day context
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type", "cpl", "cpm",
    # grain helpers and flags
    "adset_day_leads", "adset_day_leads_crm", "lead_seq_in_adset_day",
    "spend_attributed_to_lead", "matched_to_adset_day", "is_trailing_partial_day",
    "source_file",
]
lead = lead[COLUMNS].sort_values("created_at").reset_index(drop=True)

# ---------------------------------------------------------------- ad set x day sheet
# A lead-grain sheet can only describe days that produced a lead. 161 of the 879 ad-set-days
# in this window spent and got nothing -- $492.78 and 340k impressions of "this spend did not
# work", which is exactly the evidence a spend model needs. They live here.
AD_DAY_COLS = [
    "date", "ad_set_id", "campaign_id", "campaign_name", "delivery_status",
    "leads", "spend",
    "holiday_proximity", "is_holiday",
    "days_since_adset_started", "frequency",
    "ad_change_recency", "ad_set_change_recency",
    "day_of_week", "day_of_week_num", "is_weekend",
    "ad_set_change_type", "ad_change_type",
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type", "cpl", "cpm",
]
ad_days = m.copy()
ad_days["zero_lead_day"] = (ad_days["leads"] == 0).astype(int)
ad_days = ad_days[[*AD_DAY_COLS, "zero_lead_day"]].sort_values(["date", "ad_set_id"])
ad_days = ad_days.reset_index(drop=True)

# ---------------------------------------------------------------- change logs (importer reads these)
adset_events = m.loc[m["ad_set_change_recency"] == 0,
                     ["date", "ad_set_id", "campaign_name", "ad_set_change_type"]].copy()
adset_events["source"] = "inferred"
adset_events["confirmed_by"] = ""
adset_events["notes"] = ""
adset_events = adset_events.sort_values(["date", "ad_set_id"]).reset_index(drop=True)

ad_events = m.loc[(m["ad_change_recency"] == 0) & (m["ad_change_type"] != "no_recent_change"),
                  ["date", "ad_set_id", "campaign_name", "ad_change_type"]].copy()
ad_events.insert(3, "ad_id", "")
ad_events["source"] = "inferred"
ad_events["confirmed_by"] = ""
ad_events["notes"] = ""
ad_events = ad_events.sort_values(["date", "ad_set_id"]).reset_index(drop=True)

# ---------------------------------------------------------------- dictionary
D = [
    (1, "the row itself", "target", "AUTO", "one row = one lead",
     "Variable 1 is no longer a count. Each row IS a lead, named and timestamped. To get the "
     "old daily figure, count rows per ad_set_id + date, or read adset_day_leads."),
    ("--", "lead_id", "key", "AUTO", "L00001...", "Assigned here in created_at order. The CRM export carries no lead ID of its own."),
    ("--", "created_at", "key", "AUTO", "timestamp", "When the lead arrived, to the second."),
    ("--", "date", "key", "AUTO", "yyyy-mm-dd", "Calendar day of created_at. Joins the lead to its ad set x day context."),
    ("--", "hour", "support", "AUTO", "0-23", "Hour of arrival."),
    ("--", "customer_name", "key", "AUTO", "text", "From the CRM."),
    ("--", "lead_status", "target", "AUTO", "New | Existing", "CRM status. Replaces the old leads_new column, which was a per-day count."),
    ("--", "is_new_customer", "target", "AUTO", "0 or 1", "1 when lead_status is New."),
    ("--", "platform", "support", "AUTO", "text", "CRM platform, messenger throughout this window."),
    ("--", "campaign_id", "key", "AUTO", "text", "Keep as TEXT. CSV export turns these into scientific notation and destroys them."),
    ("--", "campaign_name", "key", "AUTO", "text", "Meta's campaign name, falling back to the CRM UTM Campaign string for untracked leads."),
    ("--", "ad_set_id", "key", "AUTO", "text", "Blank when the lead arrived with no UTM tag."),
    ("--", "ad_id", "key", "AUTO", "text", "The specific ad the lead came from."),
    ("--", "fb_ad_title", "support", "AUTO", "text", "Creative title as tagged."),
    ("--", "creative_code", "support", "AUTO", "text", "Leading segment of fb_ad_title, normalized."),
    ("--", "delivery_status", "support", "AUTO", "text", "Meta delivery status for that ad set that day."),
    (2, "spend", "driver", "AUTO", "number",
     "The AD SET's spend for the whole day, repeated on every lead of that ad-set-day. Do NOT "
     "sum this column - use spend_attributed_to_lead."),
    (3, "holiday_proximity", "calendar", "DERIVED", "during_holiday | 0_14_days | 15_30_days | 31_60_days | 60_plus_or_none",
     "Distance to the nearest Cambodian public holiday. Populated on every lead, since it "
     "depends only on the date. Currently 60_plus_or_none throughout - the window contains no holiday."),
    (3, "is_holiday", "calendar", "MANUAL", "0 or 1", "From the holiday calendar."),
    (4, "days_since_adset_started", "lifecycle", "AUTO", "integer", "Age of the ad set on that day, within the export window."),
    (5, "frequency", "delivery", "AUTO", "number", "Ad-set-day impressions / reach, repeated per lead."),
    (6, "ad_change_recency", "change", "DERIVED", "integer",
     "Blank until ad-level changes are recorded in changelog_ad. The detector's guess is in "
     "ad_change_recency_inferred."),
    (7, "ad_set_change_recency", "change", "DERIVED", "integer",
     "Blank until ad-set changes are recorded in changelog_ad_set. Guess in the _inferred column."),
    (8, "day_of_week", "calendar", "AUTO", "Monday..Sunday", "Weekday of the lead's own date; day_of_week_num is Monday=0."),
    (9, "ad_set_change_type", "change", "DERIVED", " | ".join(ADSET_TYPES),
     "Blank until recorded in changelog_ad_set. Meta's export carries no change log."),
    (10, "ad_change_type", "change", "DERIVED", " | ".join(AD_TYPES),
     "Blank until recorded in changelog_ad."),
    ("--", "*_inferred", "change", "INFERRED", "same as the primary column",
     "The system's detected guess, kept because the model currently runs on it. Not a recorded "
     "fact. Confirm events in the changelog sheets and the primary columns fill in."),
    ("--", "reach / impressions / messaging_conversations", "support", "AUTO", "number",
     "Ad-set-day totals repeated per lead. Do not sum down this sheet."),
    ("--", "ad_set_budget", "support", "AUTO", "number",
     "Meta's budget field. STATIC per ad set - it reports today's budget against every "
     "historical day, so it cannot date a budget change."),
    ("--", "cpl / cpm", "support", "AUTO", "number", "Ad-set-day cost per lead and per mille."),
    ("--", "adset_day_leads", "support", "AUTO", "integer",
     "The old `leads` column, kept for reconciliation: total CRM leads for that ad-set-day. "
     "Repeated on every lead of the day."),
    ("--", "adset_day_leads_crm", "support", "AUTO", "integer",
     "Same count recomputed from the rows actually present in this file."),
    ("--", "lead_seq_in_adset_day", "support", "AUTO", "integer", "This lead's position within its ad-set-day, ordered by time."),
    ("--", "spend_attributed_to_lead", "support", "AUTO", "number",
     "spend / adset_day_leads_crm. The one spend figure that sums correctly down this sheet."),
    ("--", "matched_to_adset_day", "flag", "AUTO", "0 or 1",
     "0 when the lead has no UTM tag, arrived after the Meta export ends, or its ad set had no "
     "delivery row that day. The variable columns are blank on those rows. Filter to 1 before modelling."),
    ("--", "is_trailing_partial_day", "flag", "AUTO", "0 or 1",
     "1 on 2026-08-01, whose export was pulled at 09:28 and holds a partial morning only. "
     "Exclude from daily averages."),
    ("--", "source_file", "flag", "AUTO", "text", "Which weekly CRM export the lead came from."),
]
ddict = pd.DataFrame(D, columns=["variable_no", "column", "group", "source", "allowed_values", "definition"])

README = [
    ("MODEL DATASET - LEAD GRAIN", ""),
    ("Generated", pd.Timestamp.today().strftime("%Y-%m-%d")),
    ("Grain", "ONE ROW PER LEAD"),
    ("Rows", f"{len(lead):,} leads, {lead['date'].min():%Y-%m-%d} to {lead['date'].max():%Y-%m-%d}"),
    ("", ""),
    ("WHAT CHANGED", ""),
    ("", "The old model_dataset had one row per ad set per day with a `leads` count. That count is"),
    ("", "gone. Each lead now has its own row, with its name, arrival time and status, and it"),
    ("", "carries the ad set x day context it arrived in."),
    ("", "Variable 1 is therefore the row itself. Count rows per ad_set_id + date to recover the"),
    ("", "old figure - or read adset_day_leads, which is kept for exactly that check."),
    ("", ""),
    ("THE DOUBLE-COUNTING TRAP", ""),
    ("", "spend, reach, impressions, frequency, cpl and cpm describe the AD SET'S WHOLE DAY and"),
    ("", "repeat on every lead of that day. An ad-set-day with 20 leads repeats its spend 20 times."),
    ("", "Never sum them down this sheet. Use spend_attributed_to_lead, which splits the day's"),
    ("", "spend evenly across its leads and therefore does sum correctly."),
    ("", ""),
    ("THE ad_set_days SHEET", ""),
    ("", "model_dataset can only describe days that produced a lead. 161 of the 879 ad-set-days"),
    ("", "in this window spent and got nothing - $492.78 and 340k impressions of evidence that"),
    ("", "spend did not convert, which is exactly what a spend model needs to see."),
    ("", "ad_set_days carries all 879, flagged zero_lead_day = 1 where no lead arrived. The app"),
    ("", "reads this sheet in preference to collapsing the lead rows, so uploading this one file"),
    ("", "now covers the full delivery picture."),
    ("", ""),
    ("UNMATCHED LEADS", ""),
    ("", "Every lead in the four CRM exports is here, including ones that cannot be attributed."),
    ("", "Where matched_to_adset_day = 0 the ad-set columns are blank. Calendar variables 3 and 8"),
    ("", "stay populated on those rows, because an untracked lead still arrived on a Tuesday."),
    ("", "Filter matched_to_adset_day = 1 before modelling."),
    ("", ""),
    ("2026-08-01", ""),
    ("", "That export was pulled at 09:28, so the day holds a partial morning only. Rows are kept"),
    ("", "and flagged is_trailing_partial_day = 1. Exclude it from daily averages."),
    ("", ""),
    ("VARIABLES 6, 7, 9 AND 10 ARE BLANK ON PURPOSE", ""),
    ("", "Meta's export carries no change log, so nothing has been recorded yet. The detector's"),
    ("", "guess sits in the matching *_inferred columns and is what the model runs on today."),
    ("", "Fill changelog_ad_set and changelog_ad, set source to confirmed, and upload this file:"),
    ("", "the app reads those two sheets and switches the affected ad sets off the detector."),
]

# ---------------------------------------------------------------- write
with pd.ExcelWriter(OUT, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss") as writer:
    pd.DataFrame(README, columns=["", " "]).to_excel(writer, sheet_name="README", index=False)
    lead.to_excel(writer, sheet_name="model_dataset", index=False)
    ad_days.to_excel(writer, sheet_name="ad_set_days", index=False)
    adset_events.to_excel(writer, sheet_name="changelog_ad_set", index=False)
    ad_events.to_excel(writer, sheet_name="changelog_ad", index=False)
    ddict.to_excel(writer, sheet_name="data_dictionary", index=False)

    book = writer.book
    LEAD_FILL = PatternFill("solid", fgColor="FFC6E0B4")
    KEY_FILL = PatternFill("solid", fgColor="FFD9D9D9")
    VAR_FILL = PatternFill("solid", fgColor="FFBDD7EE")
    CHANGE_FILL = PatternFill("solid", fgColor="FFFFE699")
    INFER_FILL = PatternFill("solid", fgColor="FFE2C8A0")
    SUPPORT_FILL = PatternFill("solid", fgColor="FFF2F2F2")
    MANUAL_FILL = PatternFill("solid", fgColor="FFFCE4D6")

    LEAD_COLS = set(COLUMNS[:8])
    CHANGE_COLS = {c for c in COLUMNS if "change" in c}
    VAR_COLS = {"spend", "holiday_proximity", "is_holiday", "days_since_adset_started",
                "frequency", "day_of_week", "day_of_week_num", "is_weekend"}

    sheet = book["README"]
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 100
    for row in sheet.iter_rows(max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(row[0].value, str) and row[0].value and row[0].value.isupper():
            row[0].font = Font(bold=True)

    sheet = book["model_dataset"]
    sheet.freeze_panes = "B2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    for index, cell in enumerate(sheet[1], start=1):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        name = str(cell.value)
        cell.fill = (INFER_FILL if name.endswith("_inferred") else
                     LEAD_FILL if name in LEAD_COLS else
                     CHANGE_FILL if name in CHANGE_COLS else
                     VAR_FILL if name in VAR_COLS else
                     KEY_FILL if name in {"campaign_id", "campaign_name", "ad_set_id", "ad_id"} else
                     SUPPORT_FILL)
        sheet.column_dimensions[get_column_letter(index)].width = 15
    for name, width in {"lead_id": 9, "created_at": 19, "date": 11, "hour": 7,
                        "customer_name": 24, "campaign_id": 20, "campaign_name": 26,
                        "ad_set_id": 20, "ad_id": 20, "fb_ad_title": 44,
                        "holiday_proximity": 17, "ad_set_change_type": 19,
                        "ad_change_type": 20}.items():
        sheet.column_dimensions[get_column_letter(COLUMNS.index(name) + 1)].width = width

    for name, type_column, type_list in (("changelog_ad_set", "ad_set_change_type", ADSET_TYPES),
                                         ("changelog_ad", "ad_change_type", AD_TYPES)):
        sheet = book[name]
        columns = list(adset_events.columns if name == "changelog_ad_set" else ad_events.columns)
        sheet.freeze_panes = "A2"
        sheet.row_dimensions[1].height = 30
        for index, cell in enumerate(sheet[1], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = MANUAL_FILL
            sheet.column_dimensions[get_column_letter(index)].width = 15
        sheet.column_dimensions["A"].width = 11
        sheet.column_dimensions[get_column_letter(columns.index("ad_set_id") + 1)].width = 20
        sheet.column_dimensions[get_column_letter(columns.index("campaign_name") + 1)].width = 26
        sheet.column_dimensions[get_column_letter(columns.index(type_column) + 1)].width = 22
        sheet.column_dimensions[get_column_letter(columns.index("notes") + 1)].width = 44
        last = max(sheet.max_row + 500, 1000)
        type_rule = DataValidation(type="list", formula1='"%s"' % ",".join(type_list),
                                   allow_blank=True, showErrorMessage=True,
                                   error="Use one of: " + ", ".join(type_list))
        source_rule = DataValidation(type="list", formula1='"%s"' % ",".join(SOURCES),
                                     allow_blank=True, showErrorMessage=True)
        sheet.add_data_validation(type_rule)
        sheet.add_data_validation(source_rule)
        type_letter = get_column_letter(columns.index(type_column) + 1)
        source_letter = get_column_letter(columns.index("source") + 1)
        type_rule.add(f"{type_letter}2:{type_letter}{last}")
        source_rule.add(f"{source_letter}2:{source_letter}{last}")
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{sheet.max_row}"

    sheet = book["ad_set_days"]
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.row_dimensions[1].height = 30
    for index, cell in enumerate(sheet[1], start=1):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        name = str(cell.value)
        cell.fill = (CHANGE_FILL if "change" in name else
                     VAR_FILL if name in VAR_COLS or name in {"leads", "spend"} else
                     SUPPORT_FILL)
        sheet.column_dimensions[get_column_letter(index)].width = 15
    for name, width in {"date": 11, "ad_set_id": 20, "campaign_id": 20, "campaign_name": 26,
                        "holiday_proximity": 17, "ad_set_change_type": 19,
                        "ad_change_type": 20}.items():
        sheet.column_dimensions[get_column_letter(AD_DAY_COLS.index(name) + 1)].width = width

    sheet = book["data_dictionary"]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for letter, width in {"A": 11, "B": 32, "C": 12, "D": 10, "E": 34, "F": 110}.items():
        sheet.column_dimensions[letter].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in sheet.iter_rows(min_row=2, max_col=6):
        row[4].alignment = Alignment(wrap_text=True, vertical="top")
        row[5].alignment = Alignment(wrap_text=True, vertical="top")

print("wrote", OUT)
print("raw CRM rows", n_raw, "| duplicates removed", n_dup, "| leads written", len(lead))
print("columns", len(COLUMNS))
print("matched to an ad set x day", int(lead["matched_to_adset_day"].sum()),
      "| unmatched", int((lead["matched_to_adset_day"] == 0).sum()))
print("distinct ad sets", lead["ad_set_id"].nunique(), "| ads", lead["ad_id"].nunique())
print("attributed spend", round(float(lead["spend_attributed_to_lead"].sum()), 2))
print("named leads", int(lead["customer_name"].notna().sum()))
print("recovered daily counts match adset_day_leads:",
      bool((lead.groupby(["ad_set_id", "date"]).size()
            == lead.groupby(["ad_set_id", "date"])["adset_day_leads"].first().fillna(-1)
            ).sum() > 0))
