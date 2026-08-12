"""Rebuild "Dataset 101 (all days)" fresh from the raw weekly exports + current holiday fix.

Sakda's working copy at `Dataset/Datsaa/Dataset Template/Dataset 101 (all days).xlsx` was
generated 2026-08-03 15:25, before the holiday-bucketing fix (applied 2026-08-04) and its IDs
were corrupted to floats at some point after generation (Excel auto-converting the text ID
columns to numbers, losing precision on 18-digit IDs).

This script re-derives the same two sheets, in the same column layout, from:
  - the 4 weekly CRM exports (Dataset/Datsaa/Dataset *.xlsx) -- verified byte-identical to the
    combined "Dataset 06-06 - Current.xlsx" traffic sheet for the 06-06..08-01 window
  - Dataset/master_dataset.xlsx -- already rebuilt post holiday-fix (confirmed 15
    during_holiday rows, matches the fixed calendar)
  - Dataset/Dataset/holiday_proximity.xlsx -- the fixed calendar

Output is NOT written over Sakda's working file. It goes to a new file alongside it.

Change columns (ad_set_change_type/recency, ad_change_type/recency) are blanked per the
established confirmed-vs-inferred policy (blank_policy.py) on BOTH sheets, with the detector's
guess kept in a trailing `<col>_inferred` column -- Dataset 101's ad_set_days sheet previously
had these blank with no inferred companion at all, silently dropping the guess.
"""

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CRM = [
    r"Dataset/Datsaa/Dataset 06-06 to 07-11.xlsx",
    r"Dataset/Datsaa/Dataset 07-11 to 07-18.xlsx",
    r"Dataset/Datsaa/Dataset 07-18 to 07-25.xlsx",
    r"Dataset/Datsaa/Dataset 07-25 to 08-01.xlsx",
]
MASTER = r"Dataset/master_dataset.xlsx"
HOL = r"Dataset/Dataset/holiday_proximity.xlsx"
OUT = r"Dataset/Datsaa/Dataset Template/Dataset 101 (all days) - rebuilt 2026-08-04.xlsx"

ID_COLS = ["UTM Campaign ID", "UTM Ad Set ID", "UTM Ad ID"]

# ---------------------------------------------------------------- leads
frames = [pd.read_excel(p, sheet_name=0, dtype={c: str for c in ID_COLS}) for p in CRM]
raw = pd.concat(frames, ignore_index=True)
n_raw = len(raw)

raw["created_at"] = pd.to_datetime(raw["Created At"], format="%m/%d/%Y, %I:%M:%S %p")
raw["date"] = raw["created_at"].dt.normalize()
n_dup = int(raw.duplicated(["created_at", "Customer Name", "UTM Ad ID"], keep="first").sum())
raw = raw[~raw.duplicated(["created_at", "Customer Name", "UTM Ad ID"], keep="first")]
raw = raw.sort_values("created_at").reset_index(drop=True)
raw.insert(0, "lead_id", ["L%05d" % (i + 1) for i in range(len(raw))])

lead = raw.rename(columns={
    "Platform": "platform", "Status": "lead_status", "Customer Name": "customer_name",
    "UTM Campaign": "utm_campaign", "UTM Campaign ID": "campaign_id",
    "UTM Ad Set ID": "ad_set_id", "UTM Ad ID": "ad_id", "FB Ad Title": "fb_ad_title",
})
lead["is_new_customer"] = lead["lead_status"].eq("New").astype(int)

# ---------------------------------------------------------------- ad set x day context
m = pd.read_excel(MASTER, sheet_name="master_adset_daily",
                  dtype={"ad_set_id": str, "campaign_id": str})
m["date"] = pd.to_datetime(m["date"]).dt.normalize()
m["ad_set_change_source"] = "inferred"   # no confirmed change log exists yet
m["ad_change_source"] = "inferred"

CTX = [
    "date", "ad_set_id", "campaign_id", "campaign_name", "delivery_status",
    "leads", "spend", "days_since_adset_started", "frequency",
    "ad_change_recency", "ad_set_change_recency",
    "ad_set_change_type", "ad_set_change_source",
    "ad_change_type", "ad_change_source",
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type", "cpl", "cpm",
]
ctx = m[CTX].rename(columns={"leads": "adset_day_leads", "campaign_id": "campaign_id_meta",
                             "campaign_name": "campaign_name_meta"})
lead = lead.merge(ctx, on=["ad_set_id", "date"], how="left")
lead["matched_to_adset_day"] = lead["campaign_name_meta"].notna()
lead["campaign_name"] = lead["campaign_name_meta"].fillna(lead["utm_campaign"]).fillna("")

hol = pd.read_excel(HOL)
hol["date"] = pd.to_datetime(hol["date"]).dt.normalize()
lead = lead.merge(hol[["date", "is_holiday", "holiday_proximity"]], on="date", how="left")
lead["is_holiday"] = lead["is_holiday"].fillna(0).astype(int)
lead["day_of_week"] = lead["date"].dt.day_name()
lead["day_of_week_num"] = lead["date"].dt.dayofweek
lead["is_weekend"] = lead["day_of_week_num"].isin([5, 6]).astype(int)

# ---------------------------------------------------------------- blank-per-policy (model_dataset)
def blank_change_cols(df, type_col, recency_col, source_col):
    inferred = df[source_col].eq("inferred")
    df[type_col + "_inferred"] = df[type_col]
    df[recency_col + "_inferred"] = df[recency_col]
    df.loc[inferred, type_col] = pd.NA
    df.loc[inferred, recency_col] = pd.NA
    return df


lead = blank_change_cols(lead, "ad_set_change_type", "ad_set_change_recency", "ad_set_change_source")
lead = blank_change_cols(lead, "ad_change_type", "ad_change_recency", "ad_change_source")

MODEL_COLUMNS = [
    "lead_id", "created_at", "customer_name", "lead_status", "is_new_customer", "platform",
    "campaign_id", "campaign_name", "ad_set_id", "ad_id", "fb_ad_title", "delivery_status",
    "spend", "holiday_proximity", "is_holiday", "days_since_adset_started", "frequency",
    "ad_change_recency", "ad_set_change_recency",
    "day_of_week", "day_of_week_num", "is_weekend",
    "ad_set_change_type", "ad_change_type",
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type", "cpl", "cpm", "adset_day_leads",
]
INFERRED_APPEND = [
    "ad_set_change_type_inferred", "ad_set_change_recency_inferred",
    "ad_change_type_inferred", "ad_change_recency_inferred",
]
lead_out = lead[MODEL_COLUMNS + INFERRED_APPEND].sort_values("created_at").reset_index(drop=True)

# ---------------------------------------------------------------- ad_set_days sheet
ad_days = m.copy()
ad_days["zero_lead_day"] = (ad_days["leads"] == 0).astype(int)
ad_days = blank_change_cols(ad_days, "ad_set_change_type", "ad_set_change_recency", "ad_set_change_source")
ad_days = blank_change_cols(ad_days, "ad_change_type", "ad_change_recency", "ad_change_source")

AD_DAY_COLUMNS = [
    "date", "ad_set_id", "campaign_id", "campaign_name", "delivery_status",
    "leads", "spend", "holiday_proximity", "is_holiday",
    "days_since_adset_started", "frequency",
    "ad_change_recency", "ad_set_change_recency",
    "day_of_week", "day_of_week_num", "is_weekend",
    "ad_set_change_type", "ad_change_type",
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type", "cpl", "cpm", "zero_lead_day",
]
ad_days_out = (ad_days[AD_DAY_COLUMNS + INFERRED_APPEND]
               .sort_values(["date", "ad_set_id"]).reset_index(drop=True))

# ---------------------------------------------------------------- write
with pd.ExcelWriter(OUT, engine="openpyxl", datetime_format="yyyy-mm-dd hh:mm:ss") as writer:
    lead_out.to_excel(writer, sheet_name="model_dataset", index=False)
    ad_days_out.to_excel(writer, sheet_name="ad_set_days", index=False)

    book = writer.book
    KEY_FILL = PatternFill("solid", fgColor="FFD9D9D9")
    CHANGE_FILL = PatternFill("solid", fgColor="FFFFE699")
    INFER_FILL = PatternFill("solid", fgColor="FFE2C8A0")

    for sheet_name, id_cols in (("model_dataset", ["campaign_id", "ad_set_id", "ad_id"]),
                                ("ad_set_days", ["campaign_id", "ad_set_id"])):
        sheet = book[sheet_name]
        sheet.freeze_panes = "B2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.row_dimensions[1].height = 30
        cols = list((lead_out if sheet_name == "model_dataset" else ad_days_out).columns)
        for index, cell in enumerate(sheet[1], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            name = str(cell.value)
            if name.endswith("_inferred"):
                cell.fill = INFER_FILL
            elif "change" in name:
                cell.fill = CHANGE_FILL
            elif name in id_cols:
                cell.fill = KEY_FILL
            sheet.column_dimensions[get_column_letter(index)].width = 16
        for name in id_cols:
            sheet.column_dimensions[get_column_letter(cols.index(name) + 1)].width = 20

print("wrote", OUT)
print("raw CRM rows", n_raw, "| duplicates removed", n_dup, "| leads written", len(lead_out))
print("model_dataset columns", len(lead_out.columns), "| ad_set_days columns", len(ad_days_out.columns))
print("matched to an ad set x day:", int(lead["matched_to_adset_day"].sum()),
      "| unmatched:", int((~lead["matched_to_adset_day"]).sum()))
print("distinct ad sets (model_dataset):", lead_out["ad_set_id"].nunique())
print("distinct ad sets (ad_set_days):", ad_days_out["ad_set_id"].nunique())
print("zero_lead_day rows:", int(ad_days_out["zero_lead_day"].sum()))
print("holiday_proximity value counts (model_dataset):")
print(lead_out["holiday_proximity"].value_counts().to_string())
print("id sample:", lead_out[["campaign_id", "ad_set_id", "ad_id"]].iloc[0].tolist())
print("id dtypes:", lead_out[["campaign_id", "ad_set_id", "ad_id"]].dtypes.to_string())
