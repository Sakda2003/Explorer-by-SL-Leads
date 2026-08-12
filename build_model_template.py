"""Build the weekly model-dataset template workbook.

Reads the current master dataset and re-emits it as an upload template: the same
ad set x day grain carrying all 10 declared variables, plus the two change logs
and the holiday calendar that turn variables 3, 6, 7, 9 and 10 from inferred
proxies into recorded facts.

Deliberately does NOT re-run change detection -- it reads the events already
stored in Dataset/master_dataset.xlsx. The thresholds live in
build_master_dataset.py and backend/core.py; a third copy would drift.
"""

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from blank_policy import INFERRED_COLS, apply_blank_policy, order_with_inferred

MASTER = r"Dataset/master_dataset.xlsx"
HOL = r"Dataset/Dataset/holiday_proximity.xlsx"
OUT = r"Dataset/MODEL_DATASET_TEMPLATE.xlsx"

ADSET_TYPES = ["budget_change", "targeting_change", "placement_change", "bid_change"]
AD_TYPES = ["ad_added", "ad_paused", "ad_swapped", "msg_template_change", "no_recent_change"]
SOURCES = ["confirmed", "inferred"]
PROXIMITY = ["during_holiday", "0_14_days", "15_30_days", "31_60_days", "60_plus_or_none"]

# ---------------------------------------------------------------- load
m = pd.read_excel(MASTER, sheet_name="master_adset_daily", dtype={
    "ad_set_id": str, "campaign_id": str})
m["date"] = pd.to_datetime(m["date"]).dt.normalize()
hol = pd.read_excel(HOL)
hol["date"] = pd.to_datetime(hol["date"]).dt.normalize()

# ---------------------------------------------------------------- model_dataset sheet
# Column order groups the 10 declared variables together, each followed by its
# supporting columns, so the sheet reads in the same order as the spec.
m["ad_set_change_source"] = "inferred"
m["ad_change_source"] = "inferred"
m["notes"] = ""

MODEL_COLS = [
    # keys
    "date", "ad_set_id", "campaign_id", "campaign_name", "delivery_status",
    # 1
    "leads",
    # 2
    "spend",
    # 3
    "holiday_proximity", "is_holiday", "holiday_name",
    # 4
    "days_since_adset_started", "adset_start_censored",
    # 5
    "frequency",
    # 6
    "ad_change_recency",
    # 7
    "ad_set_change_recency",
    # 8
    "day_of_week", "day_of_week_num", "is_weekend",
    # 9
    "ad_set_change_type", "ad_set_change_today", "ad_set_change_source",
    # 10
    "ad_change_type", "ad_change_today", "ad_change_source",
    # supporting
    "reach", "impressions", "messaging_conversations",
    "ad_set_budget", "ad_set_budget_type",
    "leads_new", "leads_meta_reported", "cpl", "cpm", "notes",
]
model = m[MODEL_COLS].sort_values(["date", "ad_set_id"]).reset_index(drop=True)

# blank policy: the change columns are reconstructions, so the primary column is
# left empty and the guess is preserved alongside in <col>_inferred.
model = apply_blank_policy(model)
MODEL_COLS = order_with_inferred(MODEL_COLS)
model = model[MODEL_COLS]

# variable number -> the columns that carry it, used for header colouring
VAR_OF = {
    "leads": 1, "spend": 2,
    "holiday_proximity": 3, "is_holiday": 3, "holiday_name": 3,
    "days_since_adset_started": 4, "adset_start_censored": 4,
    "frequency": 5,
    "ad_change_recency": 6,
    "ad_set_change_recency": 7,
    "day_of_week": 8, "day_of_week_num": 8, "is_weekend": 8,
    "ad_set_change_type": 9, "ad_set_change_today": 9, "ad_set_change_source": 9,
    "ad_change_type": 10, "ad_change_today": 10, "ad_change_source": 10,
}

# ---------------------------------------------------------------- change logs
ads_ev = m.loc[m["ad_set_change_recency"] == 0,
               ["date", "ad_set_id", "campaign_name", "ad_set_change_type"]].copy()
ads_ev["source"] = "inferred"
ads_ev["confirmed_by"] = ""
ads_ev["notes"] = ""
ads_ev = ads_ev.sort_values(["date", "ad_set_id"]).reset_index(drop=True)

ad_ev = m.loc[(m["ad_change_recency"] == 0) & (m["ad_change_type"] != "no_recent_change"),
              ["date", "ad_set_id", "campaign_name", "ad_change_type"]].copy()
ad_ev.insert(3, "ad_id", "")
ad_ev["source"] = "inferred"
ad_ev["confirmed_by"] = ""
ad_ev["notes"] = ""
ad_ev = ad_ev.sort_values(["date", "ad_set_id"]).reset_index(drop=True)

# ---------------------------------------------------------------- holiday calendar
cal = hol[hol["date"] >= model["date"].min()][
    ["date", "day", "is_holiday", "holiday_name", "holiday_proximity"]].copy()
cal["holiday_name"] = cal["holiday_name"].fillna("")
cal["notes"] = ""
cal = cal.reset_index(drop=True)

# ---------------------------------------------------------------- data dictionary
D = [
    ("--", "date", "key", "AUTO", "yyyy-mm-dd",
     "Calendar day. One row per ad set per day the ad set was live."),
    ("--", "ad_set_id", "key", "AUTO", "text",
     "Meta Ad set ID. Keep as TEXT -- exporting to CSV turns these into scientific notation and destroys them."),
    ("--", "campaign_id", "key", "AUTO", "text", "Meta Campaign ID, text for the same reason."),
    ("--", "campaign_name", "key", "AUTO", "text", "Campaign name from the Meta export."),
    ("--", "delivery_status", "key", "AUTO", "text", "Meta delivery status on that day."),

    (1, "leads", "target", "AUTO", "integer >= 0",
     "Daily lead count from the CRM traffic export, counted by UTM Ad Set ID x Created At date. This is the ground truth the model is fitted to. Meta's own Leads column is kept separately as leads_meta_reported and is much lower -- do not swap them."),
    (2, "spend", "driver", "AUTO", "number >= 0",
     "Amount spent (USD) from the Meta ad-set export."),

    (3, "holiday_proximity", "calendar", "DERIVED", " | ".join(PROXIMITY),
     "Bucketed distance to the nearest Cambodian public holiday, derived from the holiday_calendar sheet. Currently reports as flat / unestimable because the whole 06-06 to 08-01 window sits in 60_plus_or_none. It becomes a real regressor as soon as the data window spans a holiday."),
    (3, "is_holiday", "calendar", "MANUAL", "0 or 1", "Set on the holiday_calendar sheet, joined by date."),
    (3, "holiday_name", "calendar", "MANUAL", "text", "Holiday name, blank on ordinary days."),

    (4, "days_since_adset_started", "lifecycle", "AUTO", "integer >= 0",
     "Days since the ad set's first day in the window."),
    (4, "adset_start_censored", "lifecycle", "AUTO", "0 or 1",
     "1 = the ad set was already running on the first day of the window, so days_since_adset_started is a lower bound, not the true age. Overwrite with the real launch date if you know it."),

    (5, "frequency", "delivery", "AUTO", "number", "Impressions / reach, as reported by Meta."),

    (6, "ad_change_recency", "change", "DERIVED", "integer >= 0",
     "Days since the most recent ad-level change in this ad set, computed from the changelog_ad sheet. Falls back to days_since_adset_started when no change has been recorded yet."),
    (7, "ad_set_change_recency", "change", "DERIVED", "integer >= 0",
     "Days since the most recent ad-set-level change, computed from the changelog_ad_set sheet. Same fallback rule."),

    (8, "day_of_week", "calendar", "AUTO", "Monday..Sunday", "Weekday name."),
    (8, "day_of_week_num", "calendar", "AUTO", "0-6", "Monday = 0."),
    (8, "is_weekend", "calendar", "AUTO", "0 or 1", "Saturday or Sunday."),

    (9, "ad_set_change_type", "change", "DERIVED", " | ".join(ADSET_TYPES),
     "Type of the most recent ad-set change, carried forward from the changelog_ad_set sheet. Meta's export has no change log and its Ad Set Budget column is static, so anything not recorded by hand is INFERRED from 7-day step shifts in spend / CPM / frequency."),
    (9, "ad_set_change_today", "change", "DERIVED", "0 or 1", "1 on the day the change happened."),
    (9, "ad_set_change_source", "change", "DERIVED", " | ".join(SOURCES),
     "confirmed = a human recorded this event in changelog_ad_set. inferred = detected from delivery signatures. The whole point of this template is to move rows from inferred to confirmed."),

    (10, "ad_change_type", "change", "DERIVED", " | ".join(AD_TYPES),
     "Type of the most recent ad-level change, carried forward from the changelog_ad sheet. Inferred from UTM Ad ID activity in the CRM export when not recorded by hand."),
    (10, "ad_change_today", "change", "DERIVED", "0 or 1", "1 on the day the change happened."),
    (10, "ad_change_source", "change", "DERIVED", " | ".join(SOURCES), "As ad_set_change_source."),

    ("--", "reach", "support", "AUTO", "integer", "Meta reach."),
    ("--", "impressions", "support", "AUTO", "integer", "Meta impressions."),
    ("--", "messaging_conversations", "support", "AUTO", "integer", "Messaging conversations started."),
    ("--", "ad_set_budget", "support", "AUTO", "number",
     "Meta's Ad Set Budget field. STATIC per ad set -- it reports the current budget against every historical day, so it cannot be used to date a budget change. Record budget changes in changelog_ad_set instead."),
    ("--", "ad_set_budget_type", "support", "AUTO", "Daily | Lifetime", "Budget type."),
    ("--", "leads_new", "support", "AUTO", "integer", "Subset of leads with CRM Status = New."),
    ("--", "leads_meta_reported", "support", "AUTO", "integer",
     "Meta's own Leads column. Kept for reconciliation only; it under-reports badly. Never model on this."),
    ("--", "cpl", "support", "AUTO", "number", "spend / leads, blank when leads = 0."),
    ("--", "cpm", "support", "AUTO", "number", "spend / impressions * 1000."),
    ("--", "notes", "support", "MANUAL", "free text",
     "Anything unusual about this ad set on this day: creative approval delay, account flag, tracking outage, offline promotion running. Not modelled, but it is what lets you explain an outlier six weeks later."),
]
for _c in INFERRED_COLS:
    _row = next(r for r in D if r[1] == _c)
    D.append((_row[0], _c + "_inferred", "change", "INFERRED", _row[4],
              "The system's best guess at " + _c + ", detected from 7-day step shifts in spend / "
              "CPM / frequency (ad set) or UTM Ad ID activity in the CRM (ad level). Kept here "
              "because the model currently runs on it, but it is NOT a recorded fact. Confirm or "
              "correct it in the changelog sheet and the primary column fills in."))

ddict = pd.DataFrame(D, columns=["variable_no", "column", "group", "source", "allowed_values", "definition"])

# ---------------------------------------------------------------- README
README = [
    ("MODEL DATASET TEMPLATE", ""),
    ("Generated", pd.Timestamp.today().strftime("%Y-%m-%d")),
    ("Covers", f"{model['date'].min():%Y-%m-%d} to {model['date'].max():%Y-%m-%d}, "
               f"{model['ad_set_id'].nunique()} ad sets, {len(model):,} rows"),
    ("Grain", "one row per ad set per day"),
    ("", ""),
    ("WHAT THIS FILE IS", ""),
    ("", "The forecasting model's input schema, written out in full. Every one of the 10 declared"),
    ("", "variables has a column here, pre-filled with the data currently in the system so you can"),
    ("", "see what good input looks like rather than starting from an empty header row."),
    ("", ""),
    ("WHICH SHEETS YOU ACTUALLY FILL IN", ""),
    ("model_dataset", "AUTO. Rebuilt from the two Meta/CRM exports every import. Do not hand-edit -- "
                      "edits here are overwritten. Reference only, plus the notes column."),
    ("changelog_ad_set", "MANUAL. Variables 7 and 9. One row per ad-set change you actually made."),
    ("changelog_ad", "MANUAL. Variables 6 and 10. One row per ad-level change you actually made."),
    ("holiday_calendar", "MANUAL. Variable 3. Already filled to end of 2027 -- correct it if a date is wrong."),
    ("data_dictionary", "Reference. Every column, its source, and its allowed values."),
    ("", ""),
    ("WEEKLY ROUTINE", ""),
    ("1", "Pull the two Meta/CRM exports for the week AS XLSX, never CSV -- CSV destroys the ad set IDs."),
    ("2", "Open this file and add a row to changelog_ad_set / changelog_ad for every change you made"),
    ("", "that week. Do it from memory or from Meta's own edit history while the week is still fresh."),
    ("3", "Review the rows already there with source = inferred. If the guess is right, change source to"),
    ("", "confirmed. If it is wrong, correct the type or delete the row. If a change is missing, add it."),
    ("4", "Upload all three files. The model re-derives recency and the change dummies from your log,"),
    ("", "and only falls back to guessing for ad sets and weeks you left blank."),
    ("", ""),
    ("WHY THE CHANGE LOGS MATTER MOST", ""),
    ("", "Variables 1, 2, 4, 5 and 8 are already exact -- they are read straight off the exports and"),
    ("", "the calendar, and nothing you type will improve them."),
    ("", "Variables 6, 7, 9 and 10 are currently GUESSES. Meta's export carries no change log, so the"),
    ("", "system infers changes from 7-day step shifts in spend, CPM and frequency. That has a known"),
    ("", "circularity problem: targeting_change is detected FROM frequency and then used alongside"),
    ("", "frequency as a separate regressor. Recording the real events is the single change with the"),
    ("", "largest expected effect on forecast quality."),
    ("", "placement_change in particular cannot be detected at all from the current export -- if you"),
    ("", "never log one, that level will stay empty forever."),
    ("", "Variable 3 is currently unestimable: the whole 06-06 to 08-01 window is 60_plus_or_none, so"),
    ("", "the model has never seen a holiday. It activates on its own once the window spans one."),
    ("", ""),
    ("HONEST EXPECTATION", ""),
    ("", "This raises the ceiling; it does not guarantee hitting it. Measured on the current data,"),
    ("", "~88% of daily per-ad-set forecast error is irreducible noise -- an oracle that knew the true"),
    ("", "14-day level in advance still scored WAPE 0.557 against the model's 0.630. Better change data"),
    ("", "attacks the remaining ~12%. Weekly and portfolio-level totals are already far more accurate"),
    ("", "than the daily per-ad-set figure (WAPE 0.267 vs 0.630); judge improvement at the grain you"),
    ("", "actually make decisions at."),
    ("", "The binding constraint today is history: 56 days against 14 regressors. Every week you add"),
    ("", "helps, and the change logs make each of those weeks count for more."),
]
readme = pd.DataFrame(README, columns=["", " "])

# ---------------------------------------------------------------- write
HDR = {
    "key": "FFD9D9D9", "target": "FFC6E0B4", "driver": "FFC6E0B4",
    "calendar": "FFBDD7EE", "lifecycle": "FFBDD7EE", "delivery": "FFBDD7EE",
    "change": "FFFFE699", "support": "FFF2F2F2",
}
GROUP_OF = dict(zip(ddict["column"], ddict["group"]))
MANUAL_FILL = PatternFill("solid", fgColor="FFFCE4D6")

with pd.ExcelWriter(OUT, engine="openpyxl", datetime_format="yyyy-mm-dd") as w:
    readme.to_excel(w, sheet_name="README", index=False)
    model.to_excel(w, sheet_name="model_dataset", index=False)
    ads_ev.to_excel(w, sheet_name="changelog_ad_set", index=False)
    ad_ev.to_excel(w, sheet_name="changelog_ad", index=False)
    cal.to_excel(w, sheet_name="holiday_calendar", index=False)
    ddict.to_excel(w, sheet_name="data_dictionary", index=False)

    wb = w.book

    def style(ws, widths, header_colors=None, freeze="A2"):
        ws.freeze_panes = freeze
        for i, cell in enumerate(ws[1], start=1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if header_colors:
                fill = header_colors(cell.value)
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
        for col, wdt in widths.items():
            ws.column_dimensions[col].width = wdt
        ws.row_dimensions[1].height = 30

    # README
    ws = wb["README"]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 105
    ws.freeze_panes = "A2"
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        r[1].alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(r[0].value, str) and r[0].value.isupper() and r[0].value:
            r[0].font = Font(bold=True)

    # model_dataset
    ws = wb["model_dataset"]
    style(ws, {get_column_letter(i): 15 for i in range(1, len(MODEL_COLS) + 1)},
          header_colors=lambda v: HDR.get(GROUP_OF.get(v, "support")) if not str(v).endswith("_inferred") else "FFE2C8A0")
    for name, wdt in {"date": 11, "ad_set_id": 20, "campaign_id": 20, "campaign_name": 26,
                      "holiday_proximity": 17, "days_since_adset_started": 13,
                      "ad_set_change_type": 19, "ad_change_type": 20, "notes": 40}.items():
        ws.column_dimensions[get_column_letter(MODEL_COLS.index(name) + 1)].width = wdt
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "C2"

    # changelogs -- dropdowns + orange manual-entry headers
    def logsheet(name, cols, type_col, type_list):
        ws = wb[name]
        style(ws, {get_column_letter(i + 1): 15 for i in range(len(cols))})
        for i, c in enumerate(cols, start=1):
            ws.cell(row=1, column=i).fill = MANUAL_FILL
        ws.column_dimensions["A"].width = 11
        ws.column_dimensions[get_column_letter(cols.index("ad_set_id") + 1)].width = 20
        ws.column_dimensions[get_column_letter(cols.index("campaign_name") + 1)].width = 26
        ws.column_dimensions[get_column_letter(cols.index(type_col) + 1)].width = 22
        ws.column_dimensions[get_column_letter(cols.index("notes") + 1)].width = 44
        last = max(ws.max_row + 500, 1000)
        tv = DataValidation(type="list", formula1='"%s"' % ",".join(type_list),
                            allow_blank=True, showErrorMessage=True,
                            error="Use one of: " + ", ".join(type_list))
        sv = DataValidation(type="list", formula1='"%s"' % ",".join(SOURCES),
                            allow_blank=True, showErrorMessage=True)
        ws.add_data_validation(tv)
        ws.add_data_validation(sv)
        tl = get_column_letter(cols.index(type_col) + 1)
        sl = get_column_letter(cols.index("source") + 1)
        tv.add(f"{tl}2:{tl}{last}")
        sv.add(f"{sl}2:{sl}{last}")
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
        return ws

    logsheet("changelog_ad_set", list(ads_ev.columns), "ad_set_change_type", ADSET_TYPES)
    logsheet("changelog_ad", list(ad_ev.columns), "ad_change_type", AD_TYPES)

    # holiday_calendar
    ws = wb["holiday_calendar"]
    style(ws, {"A": 12, "B": 11, "C": 11, "D": 30, "E": 18, "F": 30})
    for i in range(1, 7):
        ws.cell(row=1, column=i).fill = MANUAL_FILL
    pv = DataValidation(type="list", formula1='"%s"' % ",".join(PROXIMITY), allow_blank=True)
    ws.add_data_validation(pv)
    pv.add(f"E2:E{ws.max_row + 200}")
    ws.auto_filter.ref = ws.dimensions

    # data_dictionary
    ws = wb["data_dictionary"]
    style(ws, {"A": 11, "B": 26, "C": 12, "D": 10, "E": 34, "F": 110})
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=6):
        r[5].alignment = Alignment(wrap_text=True, vertical="top")
        r[4].alignment = Alignment(wrap_text=True, vertical="top")
        g = GROUP_OF.get(r[1].value)
        if g:
            r[1].fill = PatternFill("solid", fgColor=HDR.get(g, "FFF2F2F2"))
        if r[3].value == "MANUAL":
            r[3].fill = MANUAL_FILL

print("wrote", OUT)
print("model_dataset  ", model.shape, model["date"].min().date(), "->", model["date"].max().date())
print("changelog_ad_set", ads_ev.shape)
print(ads_ev["ad_set_change_type"].value_counts().to_string())
print("changelog_ad   ", ad_ev.shape)
print(ad_ev["ad_change_type"].value_counts().to_string())
print("holiday_calendar", cal.shape, cal["date"].min().date(), "->", cal["date"].max().date())
print("dictionary rows ", len(ddict), "| declared vars covered",
      sorted({v for v in ddict.variable_no if v != "--"}))
