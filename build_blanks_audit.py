"""Inventory every blank cell in the two exported datasets.

Splits blanks into three kinds so the fill list is actually actionable:
  MISSING    -- a real gap; go and find the value
  STRUCTURAL -- blank is the correct answer; nothing to fill
  INFERRED   -- blanked on purpose by blank_policy; the guess is in <col>_inferred
  JOIN       -- blank because the row never matched an ad set x day
"""

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from blank_policy import BLANK_CATEGORY, INFERRED_COLS, JOIN_DEPENDENT

LEAD = r"Dataset/LEAD_LEVEL_DATASET.xlsx"
TMPL = r"Dataset/MODEL_DATASET_TEMPLATE.xlsx"
OUT = r"Dataset/BLANKS_AUDIT.xlsx"

STR_IDS = {"ad_set_id": str, "ad_id": str, "campaign_id": str}

lead = pd.read_excel(LEAD, sheet_name="lead_level", dtype=STR_IDS)
model = pd.read_excel(TMPL, sheet_name="model_dataset", dtype=STR_IDS)
cl_as = pd.read_excel(TMPL, sheet_name="changelog_ad_set", dtype=STR_IDS)
cl_ad = pd.read_excel(TMPL, sheet_name="changelog_ad", dtype=STR_IDS)
cal = pd.read_excel(TMPL, sheet_name="holiday_calendar")

WHERE = {
    "ad_set_change_type": "Meta Ads Manager -> Ad set -> Edit history. Log it in changelog_ad_set.",
    "ad_set_change_recency": "Fills itself once changelog_ad_set has events.",
    "ad_set_change_today": "Fills itself once changelog_ad_set has events.",
    "ad_change_type": "Meta Ads Manager -> Ads -> Edit history. Log it in changelog_ad.",
    "ad_change_recency": "Fills itself once changelog_ad has events.",
    "ad_change_today": "Fills itself once changelog_ad has events.",
    "ad_id": "Meta Ads Manager, or the ID Lookup sheet in the CRM export.",
    "confirmed_by": "You -- initials of whoever verified the event.",
    "notes": "You -- anything unusual that day.",
    "campaign_name": "Meta ad-set export. Blank means the ad set had no delivery row that day.",
    "delivery_status": "Meta ad-set export.",
    "spend": "Meta ad-set export.",
    "frequency": "Meta ad-set export.",
    "reach": "Meta ad-set export.",
    "impressions": "Meta ad-set export.",
    "messaging_conversations": "Meta ad-set export.",
    "ad_set_budget": "Meta ad-set export.",
    "ad_set_budget_type": "Meta ad-set export.",
    "adset_day_leads": "Derived from the CRM export once the ad-set-day exists.",
    "days_since_adset_started": "Derived once the ad set appears in the Meta export.",
    "adset_start_censored": "Derived once the ad set appears in the Meta export.",
    "holiday_name": "holiday_calendar sheet -- blank is correct on ordinary days.",
    "cpl": "Undefined at 0 leads. Nothing to fill.",
    "cpm": "Undefined at 0 impressions. Nothing to fill.",
    "updated_at": "CRM. Blank means the record was never updated. Nothing to fill.",
    "minutes_to_update": "Derived from updated_at. Nothing to fill.",
    "customer_name": "CRM.",
    "utm_campaign": "CRM UTM tagging -- blank means the lead arrived untracked.",
    "campaign_id": "CRM UTM tagging -- blank means the lead arrived untracked.",
    "ad_set_id": "CRM UTM tagging -- blank means the lead arrived untracked.",
    "fb_ad_title": "CRM.",
    "holiday_proximity": "holiday_calendar sheet.",
}


# lead-grain columns whose blanks mean "this lead arrived with no UTM tag" -- that is
# a tracking gap to chase, not a consequence of the ad-set join
UNTRACKED = ["utm_campaign", "campaign_id", "ad_set_id", "ad_id"]
# undefined for a lead that has no ad set at all; nothing to fill
UNDEFINED_WITHOUT_ADSET = [
    "adset_day_leads_crm", "lead_seq_in_adset_day", "lead_share_of_adset_day"]


def classify(col, df, sheet=""):
    if col in INFERRED_COLS:
        return "INFERRED"
    if sheet == "lead_level":
        if col in UNTRACKED:
            return "MISSING"
        if col in UNDEFINED_WITHOUT_ADSET:
            return "STRUCTURAL"
        # at lead grain these are blank only on unmatched rows -- unlike model_dataset,
        # where cpl/cpm are genuinely undefined at 0 leads / 0 impressions
        if col in ("cpl", "cpm", "spend_attributed_to_lead"):
            return "JOIN"
    if col in BLANK_CATEGORY:
        return BLANK_CATEGORY[col][0]
    if "matched_to_adset_day" in df.columns and (
            col in JOIN_DEPENDENT or col.endswith("_inferred") or col.endswith("_source")):
        return "JOIN"
    return "MISSING"


def is_blank(s):
    return s.isna() | (s.astype(str).str.strip().isin(["", "nan", "NaT", "<NA>"]))


def audit(df, sheet, dataset):
    rows = []
    for col in df.columns:
        b = is_blank(df[col])
        n = int(b.sum())
        if n == 0:
            continue
        kind = classify(col, df, sheet)
        note = BLANK_CATEGORY.get(col, ("", ""))[1]
        if sheet == "lead_level" and col in UNTRACKED:
            note = "Lead arrived with no UTM tag, so it can never be attributed."
        if sheet == "lead_level" and col in UNDEFINED_WITHOUT_ADSET:
            note = "Undefined for a lead with no ad set. Nothing to fill."
        # split join-caused blanks out of everything else
        if "matched_to_adset_day" in df.columns and kind == "JOIN":
            unmatched = df["matched_to_adset_day"].eq(0)
            n_join = int((b & unmatched).sum())
            n_other = n - n_join
            if n_join and n_other:
                rows.append((dataset, sheet, col, "JOIN", n_join, round(100 * n_join / len(df), 1),
                             "Lead never matched an ad set x day row.", WHERE.get(col, "")))
                rows.append((dataset, sheet, col, kind, n_other, round(100 * n_other / len(df), 1),
                             note, WHERE.get(col, "")))
                continue
            if n_join and not n_other:
                kind, note = "JOIN", "Lead never matched an ad set x day row."
        rows.append((dataset, sheet, col, kind, n, round(100 * n / len(df), 1),
                     note, WHERE.get(col, "")))
    return rows


ROWS = []
ROWS += audit(lead, "lead_level", "LEAD_LEVEL_DATASET.xlsx")
ROWS += audit(model, "model_dataset", "MODEL_DATASET_TEMPLATE.xlsx")
ROWS += audit(cl_as, "changelog_ad_set", "MODEL_DATASET_TEMPLATE.xlsx")
ROWS += audit(cl_ad, "changelog_ad", "MODEL_DATASET_TEMPLATE.xlsx")
ROWS += audit(cal, "holiday_calendar", "MODEL_DATASET_TEMPLATE.xlsx")

summary = pd.DataFrame(ROWS, columns=[
    "dataset", "sheet", "column", "blank_kind", "blank_cells", "pct_of_rows",
    "why_blank", "where_to_get_it"])
order = {"MISSING": 0, "INFERRED": 1, "JOIN": 2, "STRUCTURAL": 3}
summary = summary.sort_values(["blank_kind", "blank_cells"],
                              key=lambda s: s.map(order) if s.name == "blank_kind" else -s)
summary = summary.reset_index(drop=True)

# ---------------------------------------------------------------- the fill list
FILL = [
    ("1", "Ad-set change events",
     "changelog_ad_set", f"{len(cl_as)} seeded rows, all source=not_recorded",
     "Confirm / correct / delete each seeded row and add the ones the detector missed. "
     "This is variables 7 and 9, and it is the highest-value gap in the whole dataset."),
    ("2", "Ad-level change events",
     "changelog_ad", f"{len(cl_ad)} seeded rows, all source=not_recorded",
     "Same routine. Variables 6 and 10. Also fill the ad_id column -- it is empty on every "
     "seeded row because the detector works off lead activity, not ad records."),
    ("3", "True ad set launch dates",
     "model_dataset", f"{int(model['adset_start_censored'].sum())} rows across the ad sets "
     "already live on 2026-06-06",
     "days_since_adset_started counts from the first day of the export window, not from the "
     "real launch. Those ad sets are older than the number says. Look the dates up once and "
     "they are fixed forever."),
    ("4", "A data window that spans a holiday",
     "holiday_calendar", "0 holidays inside 06-06 to 08-01",
     "holiday_proximity is 60_plus_or_none on all 879 rows, so variable 3 cannot be estimated "
     "at all right now. Nothing to fill -- it resolves itself as the window grows."),
    ("5", "Untracked leads",
     "lead_level", f"{int((lead['has_utm_ad_set'] == 0).sum())} leads with no UTM ad set",
     "These arrived without UTM tags, so they can never be attributed. Worth checking whether "
     "a specific entry point is dropping the tags."),
    ("6", "Leads whose ad set had no delivery row",
     "lead_level", f"{int(((lead['matched_to_adset_day'] == 0) & (lead['has_utm_ad_set'] == 1) & (lead['outside_perf_window'] == 0)).sum())} leads",
     "A lead came in attributed to an ad set that Meta reported no spend or impressions for "
     "that day. Usually a stale UTM tag on an old creative. Decide whether those ad sets "
     "belong in the dashboard."),
    ("7", "Notes",
     "both files", "empty on every row",
     "Optional, but it is what lets you explain an outlier six weeks later."),
]
fill = pd.DataFrame(FILL, columns=["priority", "what_is_missing", "where", "how_much", "what_to_do"])

# ---------------------------------------------------------------- cell-level detail
def cells(df, keycols, sheet):
    out = []
    kinds = {c: classify(c, df, sheet) for c in df.columns}
    for col in df.columns:
        if kinds[col] == "STRUCTURAL":
            continue
        b = is_blank(df[col])
        if not b.any():
            continue
        sub = df.loc[b, keycols].copy()
        sub["sheet"] = sheet
        sub["blank_column"] = col
        sub["blank_kind"] = kinds[col]
        out.append(sub)
    return pd.concat(out, ignore_index=True) if out else pd.DataFrame()


lead_cells = cells(lead, ["lead_id", "created_at", "ad_set_id"], "lead_level")
model_cells = cells(model, ["date", "ad_set_id", "campaign_name"], "model_dataset")

# per-ad-set view of what is missing, so gaps can be chased one ad set at a time
by_adset = (model.assign(
    missing_adset_change=is_blank(model["ad_set_change_type"]).astype(int),
    missing_ad_change=is_blank(model["ad_change_type"]).astype(int))
    .groupby(["ad_set_id", "campaign_name"], as_index=False)
    .agg(days=("date", "size"),
         first_day=("date", "min"), last_day=("date", "max"),
         leads=("leads", "sum"), spend=("spend", "sum"),
         start_censored=("adset_start_censored", "max"),
         days_missing_adset_change=("missing_adset_change", "sum"),
         days_missing_ad_change=("missing_ad_change", "sum"))
    .sort_values("leads", ascending=False))
by_adset["needs_real_launch_date"] = by_adset["start_censored"].map({1: "YES", 0: ""})
by_adset = by_adset.drop(columns=["start_censored"])

# ---------------------------------------------------------------- write
KIND_DOC = pd.DataFrame([
    ("MISSING", "A real gap. Someone has to go and find the value.", "Fill it."),
    ("INFERRED", "Deliberately blanked. The system's guess is preserved in the matching "
                 "<column>_inferred column, and the model still runs on that guess.",
     "Confirm the guess in the changelog sheet and the primary column fills in."),
    ("JOIN", "Blank because that lead never matched an ad set x day row -- no UTM tag, "
             "outside the Meta export window, or the ad set had no delivery that day.",
     "Fix the tagging or accept it; filter matched_to_adset_day = 1 before modelling."),
    ("STRUCTURAL", "Blank is the correct answer. There is nothing to fill.", "Ignore."),
], columns=["blank_kind", "meaning", "action"])

with pd.ExcelWriter(OUT, engine="openpyxl", datetime_format="yyyy-mm-dd") as w:
    KIND_DOC.to_excel(w, sheet_name="how_to_read_this", index=False)
    fill.to_excel(w, sheet_name="what_to_fill", index=False)
    summary.to_excel(w, sheet_name="blank_summary", index=False)
    by_adset.to_excel(w, sheet_name="gaps_by_ad_set", index=False)
    if len(lead_cells):
        lead_cells.to_excel(w, sheet_name="blank_cells_lead", index=False)
    if len(model_cells):
        model_cells.to_excel(w, sheet_name="blank_cells_model", index=False)

    wb = w.book
    KF = {"MISSING": "FFF8CBAD", "INFERRED": "FFFFE699",
          "JOIN": "FFDDEBF7", "STRUCTURAL": "FFEDEDED"}
    for name in wb.sheetnames:
        ws = wb[name]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 28
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for col in ws.columns:
            hdr = col[0].value
            width = {"why_blank": 46, "where_to_get_it": 52, "what_to_do": 78,
                     "meaning": 70, "action": 46, "column": 26, "blank_column": 26,
                     "campaign_name": 26, "ad_set_id": 20, "sheet": 17,
                     "dataset": 30, "what_is_missing": 30, "how_much": 34,
                     "created_at": 19, "lead_id": 9}.get(hdr, 14)
            ws.column_dimensions[col[0].column_letter].width = width
            if hdr in ("why_blank", "where_to_get_it", "what_to_do", "meaning", "action"):
                for c in col[1:]:
                    c.alignment = Alignment(wrap_text=True, vertical="top")
            if hdr == "blank_kind":
                for c in col[1:]:
                    if c.value in KF:
                        c.fill = PatternFill("solid", fgColor=KF[c.value])

print("wrote", OUT)
print(summary.to_string(index=False,
      columns=["sheet", "column", "blank_kind", "blank_cells", "pct_of_rows"]))
print("\nblank cells by kind:")
print(summary.groupby("blank_kind")["blank_cells"].agg(["sum", "count"]).to_string())
